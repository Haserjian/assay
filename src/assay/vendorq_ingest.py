"""Questionnaire ingestion for VendorQ."""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Dict, List

from assay.vendorq_models import (
    SCHEMA_VERSION_QUESTION,
    VendorQInputError,
    is_yes_no_question,
    now_utc_iso,
    validate_vendorq_schema,
    write_json,
)


def _normalize_question(question_id: str, question_text: str, type_hint: str, required_format: str) -> Dict[str, str]:
    qid = question_id.strip() or ""
    qtext = question_text.strip()
    if not qtext:
        raise VendorQInputError("empty_question_text")

    hint = (type_hint or "unknown").strip().lower()
    if hint not in {"yes_no", "free_text", "numeric", "multi_select", "unknown"}:
        hint = "unknown"

    req = required_format.strip() or "text"
    return {
        "question_id": qid,
        "question_text": qtext,
        "type_hint": hint,
        "required_format": req,
    }


def _infer_hint_from_text(text: str) -> str:
    if is_yes_no_question(text):
        return "yes_no"
    t = text.strip().lower()
    if any(k in t for k in ("how many", "number", "count", "%", "latency", "cost")):
        return "numeric"
    return "free_text"


def _rows_from_csv(path: Path) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise VendorQInputError("csv_missing_header")

        for i, row in enumerate(reader, start=1):
            raw_id = str(row.get("question_id") or row.get("id") or f"q{i:03d}")
            raw_text = str(row.get("question_text") or row.get("question") or row.get("text") or "")
            raw_hint = str(row.get("type_hint") or _infer_hint_from_text(raw_text))
            raw_fmt = str(row.get("required_format") or "text")
            q = _normalize_question(raw_id, raw_text, raw_hint, raw_fmt)
            if not q["question_id"]:
                q["question_id"] = f"q{i:03d}"
            out.append(q)
    return out


def _rows_from_md(path: Path) -> List[Dict[str, str]]:
    text = path.read_text(encoding="utf-8")
    out: List[Dict[str, str]] = []
    in_code = False
    n = 0
    for raw in text.splitlines():
        line = raw.strip()
        if line.startswith("```"):
            in_code = not in_code
            continue
        if in_code or not line or line.startswith("#"):
            continue

        # list items: -, *, 1.
        m = re.match(r"^(?:[-*]|\d+[.)])\s+(.*)$", line)
        if m:
            line = m.group(1).strip()

        # Support "Q12: ..." style IDs.
        qid_match = re.match(r"^(Q\d+|q\d+|[A-Za-z_][A-Za-z0-9_\-]*):\s+(.*)$", line)
        if qid_match:
            qid = qid_match.group(1)
            qtext = qid_match.group(2)
        else:
            n += 1
            qid = f"q{n:03d}"
            qtext = line

        if not qtext:
            continue
        out.append(_normalize_question(qid, qtext, _infer_hint_from_text(qtext), "text"))

    if not out:
        raise VendorQInputError("markdown_no_questions_found")
    return out


def _rows_from_xlsx(path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise VendorQInputError(f"xlsx_support_unavailable: {e}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise VendorQInputError("xlsx_empty")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    def _cell(row: Any, *names: str, default: str = "") -> str:
        for nm in names:
            if nm in col and col[nm] < len(row):
                v = row[col[nm]]
                if v is not None:
                    return str(v)
        return default

    out: List[Dict[str, str]] = []
    for i, row in enumerate(rows[1:], start=1):
        qid = _cell(row, "question_id", "id", default=f"q{i:03d}")
        qtext = _cell(row, "question_text", "question", "text", default="")
        if not qtext.strip():
            continue
        hint = _cell(row, "type_hint", default=_infer_hint_from_text(qtext))
        req = _cell(row, "required_format", default="text")
        out.append(_normalize_question(qid, qtext, hint, req))

    if not out:
        raise VendorQInputError("xlsx_no_questions_found")
    return out


def ingest_questionnaire(in_path: Path, out_path: Path, source_label: str = "") -> Dict[str, Any]:
    if not in_path.exists():
        raise VendorQInputError(f"input_not_found: {in_path}")

    suffix = in_path.suffix.lower()
    if suffix == ".csv":
        questions = _rows_from_csv(in_path)
    elif suffix in {".md", ".markdown"}:
        questions = _rows_from_md(in_path)
    elif suffix == ".xlsx":
        questions = _rows_from_xlsx(in_path)
    else:
        raise VendorQInputError(f"unsupported_input_format: {in_path.suffix}")

    if not questions:
        raise VendorQInputError("no_questions_found")

    payload: Dict[str, Any] = {
        "schema_version": SCHEMA_VERSION_QUESTION,
        "generated_at": now_utc_iso(),
        "source": source_label.strip() or in_path.name,
        "questions": questions,
    }

    validate_vendorq_schema("vendorq.question.v1.schema.json", payload)
    write_json(out_path, payload)
    return payload
